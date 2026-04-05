# Scholarship Data Text File to Excel Converter
# Converts fixed-width .txt files into a structured Excel sheet.
#
# Field Layout (based on sample line):
#   Col 1-2    : ABPS Transaction Code (e.g., "77")
#   Col 3-17   : Aadhaar Number (15 digits, strip leading zeros)
#   Col 18-47  : Beneficiary Name (30 chars, strip whitespace)
#   Col 48-56  : IIN Number (9 digits, strip leading zeros)
#   Col 57-63  : User Number (7 chars)

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# CONFIG – change only these if layout shifts
# ──────────────────────────────────────────────
INPUT_FOLDER  = r"C:\Users\Aravind\Downloads\Scholarship_Data"
OUTPUT_FILE   = r"C:\Users\Aravind\Downloads\Scholarship_Data\Scholarship_Output.xlsx"

# Fixed-width slice positions (0-indexed, end is exclusive)
SLICES = {
    "ABPS Transaction Code": (0,   2),
    "Aadhaar Number":        (2,  17),
    "Beneficiary Name":      (17, 47),
    "IIN Number":            (47, 56),
    "User Number":           (56, 63),
}

# Columns whose leading zeros should be stripped
STRIP_LEADING_ZEROS = {"Aadhaar Number", "IIN Number"}

# ──────────────────────────────────────────────
# PARSE
# ──────────────────────────────────────────────
def parse_line(line: str) -> dict | None:
    """Return a dict of fields for a data line, or None if line is blank/header."""
    if not line.strip():
        return None
    row = {}
    for field, (start, end) in SLICES.items():
        raw = line[start:end] if len(line) > start else ""
        value = raw.strip()
        if field in STRIP_LEADING_ZEROS:
            value = value.lstrip("0") or "0"
        row[field] = value
    return row


def collect_records(folder: str) -> list[dict]:
    records = []
    txt_files = [f for f in os.listdir(folder) if f.lower().endswith(".txt")]
    if not txt_files:
        raise FileNotFoundError(f"No .txt files found in: {folder}")
    for fname in txt_files:
        path = os.path.join(folder, fname)
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            for line in fh:
                record = parse_line(line.rstrip("\n"))
                if record:
                    records.append(record)
    return records


# ──────────────────────────────────────────────
# EXCEL STYLING HELPERS
# ──────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", start_color="1F3864")   # dark navy
ALT_FILL     = PatternFill("solid", start_color="EEF2FF")   # light blue
WHITE_FILL   = PatternFill("solid", start_color="FFFFFF")
BORDER_COLOR = "B8C4DE"

def thin_border():
    s = Side(style="thin", color=BORDER_COLOR)
    return Border(left=s, right=s, top=s, bottom=s)

HEADERS = list(SLICES.keys())
COL_WIDTHS = {
    "ABPS Transaction Code": 22,
    "Aadhaar Number":        18,
    "Beneficiary Name":      32,
    "IIN Number":            14,
    "User Number":           14,
}


def build_excel(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scholarship Data"

    # ── Header row ──
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()

    ws.row_dimensions[1].height = 30

    # ── Data rows ──
    for row_idx, record in enumerate(records, start=2):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = thin_border()
            cell.alignment = Alignment(
                horizontal="center" if header != "Beneficiary Name" else "left",
                vertical="center"
            )

    # ── Column widths ──
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS[header]

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Summary row at bottom ──
    summary_row = len(records) + 2
    ws.cell(row=summary_row, column=1, value=f"Total Records: {len(records)}").font = Font(
        name="Arial", bold=True, size=10, color="1F3864"
    )

    wb.save(output_path)
    print(f"✅  Excel saved → {output_path}  ({len(records)} records)")


# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
if __name__ == "__main__":
    print(f"📂  Reading .txt files from: {INPUT_FOLDER}")
    records = collect_records(INPUT_FOLDER)
    print(f"📋  Parsed {len(records)} records")
    build_excel(records, OUTPUT_FILE)
